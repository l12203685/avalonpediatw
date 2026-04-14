#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Parse Avalonpedia master Excel workbook into per-sheet YAML files.

Output consumed by Astro Content Collections under content/_data/.

Usage:
    python scripts/parse_master.py [--input PATH] [--output DIR] [--verbose]
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import yaml

# openpyxl on Python 3.14 has a pivot-cache parse bug with this workbook;
# pandas (with openpyxl as the XLSX engine) works because it skips pivot
# machinery we don't need.
import pandas as pd

try:
    import openpyxl  # noqa: F401  (optional, for merged-cell introspection)
    HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    HAS_OPENPYXL = False


TAIPEI = timezone(timedelta(hours=8))


# ---------- helpers ----------

def taipei_now_iso() -> str:
    return datetime.now(TAIPEI).strftime("%Y-%m-%d %H:%M:%S +08")


_SAFE_CHARS = re.compile(r"[^\w\u4e00-\u9fff\-]+", re.UNICODE)


def to_filename(sheet_name: str) -> str:
    """Convert sheet name to a safe filename (keep CJK + ascii)."""
    name = sheet_name.strip()
    name = _SAFE_CHARS.sub("_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name or "sheet"


def cell_is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def normalize_cell(value: Any) -> Any:
    """Make values YAML/serializer friendly."""
    if cell_is_empty(value):
        return None
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return float(value)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    if isinstance(value, str):
        return value.strip()
    return value


def unique_headers(raw: list[Any]) -> list[str | None]:
    """Turn first row into header strings; empty header -> None (skip col)."""
    seen: dict[str, int] = {}
    out: list[str | None] = []
    for v in raw:
        if cell_is_empty(v):
            out.append(None)
            continue
        h = str(v).strip()
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        out.append(h)
    return out


def row_to_dict(row: list[Any], headers: list[str | None]) -> dict[str, Any] | None:
    out: dict[str, Any] = {}
    any_value = False
    for h, v in zip(headers, row):
        if h is None:
            continue
        norm = normalize_cell(v)
        if norm is not None:
            any_value = True
        out[h] = norm
    return out if any_value else None


# ---------- core read ----------

def read_sheet_dataframe(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    """Read a sheet as a header-less DataFrame."""
    df = pd.read_excel(
        xlsx_path, sheet_name=sheet_name, header=None, engine="openpyxl",
        dtype=object,
    )
    return df


def fill_merged_ranges(xlsx_path: Path, sheet_name: str, df: pd.DataFrame,
                        log: logging.Logger) -> pd.DataFrame:
    """Use openpyxl to unmerge: propagate anchor value across each merged range.

    Gracefully degrades if openpyxl can't load the workbook (pivot bug).
    """
    if not HAS_OPENPYXL:
        return df
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        if sheet_name not in wb.sheetnames:
            return df
        ws = wb[sheet_name]
        ranges = list(ws.merged_cells.ranges)
        if not ranges:
            return df
        log.debug("  sheet %s: %d merged ranges", sheet_name, len(ranges))
        for mr in ranges:
            r0, c0, r1, c1 = mr.min_row - 1, mr.min_col - 1, mr.max_row - 1, mr.max_col - 1
            if r0 >= len(df) or c0 >= len(df.columns):
                continue
            anchor = df.iat[r0, c0]
            for rr in range(r0, min(r1 + 1, len(df))):
                for cc in range(c0, min(c1 + 1, len(df.columns))):
                    if rr == r0 and cc == c0:
                        continue
                    df.iat[rr, cc] = anchor
        return df
    except Exception as e:
        log.warning("  openpyxl merge-fill skipped for %s: %s", sheet_name, e)
        return df


def sheet_to_records(df: pd.DataFrame) -> tuple[list[dict[str, Any]], int, int]:
    """Turn DataFrame into list[dict]. Returns (records, skipped_rows, header_cols)."""
    if df.empty:
        return [], 0, 0
    headers = unique_headers(list(df.iloc[0].tolist()))
    header_cols = sum(1 for h in headers if h is not None)
    records: list[dict[str, Any]] = []
    skipped = 0
    for _, row in df.iloc[1:].iterrows():
        rec = row_to_dict(row.tolist(), headers)
        if rec is None:
            skipped += 1
            continue
        records.append(rec)
    return records, skipped, header_cols


# ---------- write ----------

def write_yaml(out_path: Path, sheet_name: str, records: list[dict[str, Any]],
               generated_at: str, source_file: str) -> int:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    meta_header = (
        f"# Generated from {source_file} @ {generated_at}\n"
        f"# Source sheet: {sheet_name}\n"
        f"# Rows: {len(records)}\n"
    )
    body = yaml.safe_dump(records, allow_unicode=True, sort_keys=False,
                          default_flow_style=False)
    text = meta_header + body
    out_path.write_text(text, encoding="utf-8")
    return len(text.encode("utf-8"))


# ---------- CLI ----------

def build_logger(verbose: bool) -> logging.Logger:
    log = logging.getLogger("parse_master")
    log.handlers.clear()
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("%(message)s"))
    log.addHandler(h)
    log.setLevel(logging.DEBUG if verbose else logging.INFO)
    return log


SPECIAL_SHEETS = {
    "角色": "roles",
    "陣容": "team_composition",
    "規則": "rules",
}


def resolve_output_name(sheet_name: str, used_names: set[str]) -> str:
    for needle, target in SPECIAL_SHEETS.items():
        if needle in sheet_name and target not in used_names:
            return target
    base = to_filename(sheet_name)
    if base in used_names:
        base = f"raw_{base}"
    return base


def find_fallback_input() -> Path | None:
    candidates = [
        Path(r"E:/阿瓦隆百科/阿瓦隆百科.xlsx"),
        Path(r"C:/Users/admin/workspace/avalonpediatw/阿瓦隆百科.xlsx"),
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", default=r"E:/阿瓦隆百科/阿瓦隆百科.xlsx")
    ap.add_argument("--output", default="content/_data/")
    ap.add_argument("--verbose", "-v", action="store_true")
    ap.add_argument("--no-merge-fill", action="store_true",
                    help="skip openpyxl merged-cell unmerge pass")
    args = ap.parse_args(argv)

    log = build_logger(args.verbose)

    xlsx = Path(args.input)
    if not xlsx.exists():
        alt = find_fallback_input()
        if alt is None:
            log.error("Input not found: %s", xlsx)
            try:
                for p in Path("E:/").iterdir():
                    log.error("  E:/ entry: %s", p.name)
            except Exception:
                pass
            return 2
        log.warning("Input %s missing; using fallback %s", xlsx, alt)
        xlsx = alt

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    log.info("Reading %s", xlsx)
    xl = pd.ExcelFile(xlsx, engine="openpyxl")
    log.info("Sheets: %d", len(xl.sheet_names))

    generated_at = taipei_now_iso()
    source_file = xlsx.name

    used_names: set[str] = set()
    summary: list[dict[str, Any]] = []

    for sheet_name in xl.sheet_names:
        entry: dict[str, Any] = {"sheet": sheet_name, "status": "ok"}
        try:
            df = xl.parse(sheet_name, header=None, dtype=object)
            if not args.no_merge_fill:
                df = fill_merged_ranges(xlsx, sheet_name, df, log)
            records, skipped, hdr_cols = sheet_to_records(df)
            out_name = resolve_output_name(sheet_name, used_names)
            used_names.add(out_name)
            out_path = out_dir / f"{out_name}.yaml"
            size = write_yaml(out_path, sheet_name, records, generated_at, source_file)
            entry.update(file=out_path.name, rows=len(records),
                         skipped=skipped, header_cols=hdr_cols, bytes=size)
            log.info("  [%s] -> %s rows=%d skipped=%d bytes=%d",
                     sheet_name, out_path.name, len(records), skipped, size)
        except Exception as e:
            entry.update(status="error", error=str(e))
            log.error("  [%s] FAILED: %s", sheet_name, e)
            log.debug(traceback.format_exc())
        summary.append(entry)

    ok = sum(1 for s in summary if s["status"] == "ok")
    err = sum(1 for s in summary if s["status"] != "ok")
    total_rows = sum(s.get("rows", 0) for s in summary)
    total_bytes = sum(s.get("bytes", 0) for s in summary)
    log.info("---")
    log.info("Summary: %d ok / %d error / %d sheets", ok, err, len(summary))
    log.info("Total rows: %d / Total bytes: %d", total_rows, total_bytes)
    log.info("Output dir: %s", out_dir.resolve())

    summary_path = out_dir / "_parse_summary.yaml"
    summary_path.write_text(
        f"# Generated @ {generated_at}\n"
        + yaml.safe_dump({
            "source": str(xlsx),
            "generated_at": generated_at,
            "sheets": summary,
        }, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )

    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
